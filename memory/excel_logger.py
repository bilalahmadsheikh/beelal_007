"""
excel_logger.py — BilalAgent v2.0 Excel Job Application Logger
Tracks job applications in applied_jobs.xlsx with proper formatting.
"""

import os
from datetime import datetime

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "memory", "applied_jobs.xlsx")


def _ensure_workbook():
    """Create the Excel workbook with headers if it doesn't exist."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if os.path.exists(EXCEL_PATH):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Headers
    headers = ["Date", "Company", "Job Title", "Score", "Status", 
               "Location", "Salary", "URL", "Source", "Cover Letter"]
    
    # Style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Column widths
    widths = [12, 25, 35, 8, 12, 20, 20, 50, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"[EXCEL] Created {EXCEL_PATH}")


def log_application(job: dict, score: dict, status: str = "applied", 
                    cover_letter_path: str = "") -> None:
    """
    Log a job application to the Excel file.
    
    Args:
        job: Job dict with title, company, etc.
        score: Score dict with score, matching_skills, etc.
        status: Application status (applied, saved, rejected, interview, offer)
        cover_letter_path: Path to saved cover letter file
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    _ensure_workbook()
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    row_data = [
        datetime.now().strftime("%Y-%m-%d"),
        job.get("company", ""),
        job.get("title", ""),
        score.get("score", 0),
        status,
        job.get("location", ""),
        job.get("salary", "Not listed"),
        job.get("url", ""),
        job.get("site_name", ""),
        cover_letter_path,
    ]
    
    row_num = ws.max_row + 1
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Color-code by score
    if score.get("score", 0) >= 80:
        score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif score.get("score", 0) >= 60:
        score_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if col_idx == 4:  # Score column
            cell.fill = score_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    wb.save(EXCEL_PATH)
    print(f"[EXCEL] Logged: {job.get('title', '')} @ {job.get('company', '')} (Score: {score.get('score', 0)})")


def get_applications(status: str = None) -> list:
    """
    Get logged applications from Excel.
    
    Args:
        status: Optional filter (applied, saved, rejected, interview, offer)
        
    Returns:
        List of dicts with application data
    """
    from openpyxl import load_workbook
    
    if not os.path.exists(EXCEL_PATH):
        return []
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    apps = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        app = dict(zip(headers, row))
        if status and app.get("Status", "").lower() != status.lower():
            continue
        apps.append(app)
    
    return apps


def display_applications(apps: list) -> str:
    """Format applications for display."""
    if not apps:
        return "No applications found."
    
    lines = [f"\n📋 Job Applications ({len(apps)} total)\n{'═' * 50}"]
    
    for app in apps:
        score = app.get("Score", 0)
        emoji = "🟢" if score >= 80 else "🟡" if score >= 60 else "🔴"
        lines.append(f"\n{emoji} {app.get('Job Title', '')} @ {app.get('Company', '')}")
        lines.append(f"   Score: {score}/100 | Status: {app.get('Status', '')}")
        lines.append(f"   {app.get('Location', '')} | {app.get('Salary', '')}")
        lines.append(f"   Applied: {app.get('Date', '')} | Source: {app.get('Source', '')}")
    
    lines.append(f"\n{'═' * 50}")
    lines.append(f"📁 Full data: {EXCEL_PATH}")
    return "\n".join(lines)


# ─────────────────────────────────────────────────
# Phase 5: Gig tracking (gigs_created.xlsx)
# ─────────────────────────────────────────────────

GIGS_EXCEL_PATH = os.path.join(PROJECT_ROOT, "memory", "gigs_created.xlsx")


def _ensure_gigs_workbook():
    """Create the gigs Excel workbook with headers if it doesn't exist."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if os.path.exists(GIGS_EXCEL_PATH):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Gigs"
    
    headers = ["Date", "Platform", "Service", "Title", "Status", "URL", "Price Range"]
    
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    widths = [12, 12, 18, 50, 12, 50, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    ws.freeze_panes = "A2"
    
    os.makedirs(os.path.dirname(GIGS_EXCEL_PATH), exist_ok=True)
    wb.save(GIGS_EXCEL_PATH)
    print(f"[EXCEL] Created {GIGS_EXCEL_PATH}")


def log_gig(platform: str, service: str, title: str, status: str = "draft",
            url: str = "", price: str = "") -> None:
    """
    Log a gig to gigs_created.xlsx.
    
    Args:
        platform: fiverr, upwork, freelancer
        service: mlops, chatbot, blockchain, data_science, backend
        title: Gig title
        status: draft, published, active, paused
        url: Published gig URL
        price: Price range string
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side, PatternFill
    
    _ensure_gigs_workbook()
    
    wb = load_workbook(GIGS_EXCEL_PATH)
    ws = wb.active
    
    row_data = [
        datetime.now().strftime("%Y-%m-%d"),
        platform,
        service,
        title,
        status,
        url,
        price,
    ]
    
    row_num = ws.max_row + 1
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    
    # Color by status
    status_colors = {
        "draft": PatternFill(start_color="FFEB9C", fill_type="solid"),
        "published": PatternFill(start_color="C6EFCE", fill_type="solid"),
        "active": PatternFill(start_color="C6EFCE", fill_type="solid"),
        "paused": PatternFill(start_color="FFC7CE", fill_type="solid"),
    }
    status_fill = status_colors.get(status.lower(), PatternFill())
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if col_idx == 5:  # Status column
            cell.fill = status_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    wb.save(GIGS_EXCEL_PATH)
    print(f"[EXCEL] Logged gig: {title} ({platform}/{service}) — {status}")


def get_gigs(status: str = None) -> list:
    """Get logged gigs from Excel."""
    from openpyxl import load_workbook
    
    if not os.path.exists(GIGS_EXCEL_PATH):
        return []
    
    wb = load_workbook(GIGS_EXCEL_PATH)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    gigs = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        gig = dict(zip(headers, row))
        if status and gig.get("Status", "").lower() != status.lower():
            continue
        gigs.append(gig)
    
    return gigs


if __name__ == "__main__":
    print("=" * 50)
    print("Excel Logger Test")
    print("=" * 50)
    
    # Test logging
    test_job = {"title": "AI Engineer", "company": "TestCorp", 
                "location": "Remote", "salary": "$120k", "url": "https://example.com",
                "site_name": "indeed"}
    test_score = {"score": 85, "matching_skills": ["Python", "MLOps"], "missing": ["Go"]}
    
    log_application(test_job, test_score, "saved")
    
    apps = get_applications()
    print(display_applications(apps))
    
    # Test gig logging
    log_gig("fiverr", "mlops", "I will build MLOps pipelines", "draft", price="$20-$130")
    print(f"\nGigs: {len(get_gigs())} total")

